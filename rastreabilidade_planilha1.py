import streamlit as st
from streamlit import caching
import pandas as pd
from io import StringIO
import numpy as np
from st_aggrid import GridOptionsBuilder, AgGrid, GridUpdateMode, DataReturnMode, JsCode
import base64
import json
import smtplib
import time
import datetime
from datetime import date, datetime, time
import pytz
from io import BytesIO
from openpyxl import load_workbook, Workbook
from google.cloud import firestore
from google.oauth2 import service_account

###############################################################################
# Configurações da página
###############################################################################

st.set_page_config(page_title="Rastreabilidade",layout="wide")

###############################################################################
# Configurando acesso ao firebase
###############################################################################
textkey_2 = """{\n  \"type\": \"service_account\",\n  \"project_id\": \"lid-rastr\",\n  \"private_key_id\": \"a7c22fdbd57a9b70915020a3075fe968298d9b07\",\n  \"private_key\": \"-----BEGIN PRIVATE KEY-----\\nMIIEvQIBADANBgkqhkiG9w0BAQEFAASCBKcwggSjAgEAAoIBAQDGij/1qDDBibyf\\nooQJBsT+afWbqJchuhDgLixrMMrBknCzDKqnMKQGPAq3d3E5A95vf0tyJOQflefW\\ntb89OApzGmx9fDzN4NKIgUPQZJYB1GG/x+JeyQqJCVPzpG3R5t6WtK7XSWxg5WBx\\nQ1JS2KAYKZehZzhp1zDr25w5LeQlNpIbKpAS67yOO4nieA3ft5XGF/YCTPBA5IrV\\nkPusIUE0nbFGPIqHnq6Cj3pnVzpd/iweec9UiGVeUCW3Dw7kCJQom1lHoVYksMhx\\nZSQGzAiIOu6s4OLE/CPXpoa6P8i2erppcQyYIXJ6dL2W4AJ/CKKAt0zRN1N13pgt\\nnhpUIXELAgMBAAECggEAQ8JlR9MKKNN4Y8cUZvw/eVDyeRiV0/Xr0ocPs9moKV5w\\nRjt5dqwcHuCZC7qhEsNmRAle12sNzFXeFSJcTWl174jCJCWlnuIvGFV9rn7Vz3QL\\nlGeEs7LLfK+JTmr87BluOGMcFO/DJGLEgoNmck3qfbScQoK29zBxSt3duIoYBja5\\n6MmkTPPnQhi/gWiI7V/vsAcLvSQH1+PqYP/OzzIsd/Cp9vwNx/u+i50AndsXjDUR\\n2HcfwOIeE3h4H3Zh1Fvzgsznm2fjKEOJAG6wzJtzbe0mBRslld2v+Wuy3QuDM1+N\\nDA1hvSatnxaLNfGCnA70YmDoe+ueoVcGdUl2u0NbYQKBgQD7F/GUeV6Zp/ouh7Ka\\nzmtAtJDk5nWy43ujHH/E+0v5jqz/W0RbvAYuK321IyN1wBA/LHqHMit3Z+XxEw9Z\\ndKGEASNY1Lv7HEgoks6Chn02+k9HJLYD0sdQgIzhMKrIP5UOmwDQT+BNv+nWKXhn\\nyrmwyPnNZ8M2e2+hJBt6dceg2wKBgQDKa2pebKAq7DT0XeU6x6vX/kGtBCycgvzn\\nKb7R9Z7QnUk0IAfAUtQ0wwclhi1R08XpZwzL3BEO6EAo5fzJa/6ObEv9SUwsx7YP\\nriekUDkGqxacGFIe7QqVEHxQgDQ0eUrGd6SOELGNOmi5etVZJlsMZKp6GmvyQL8n\\nbMQUqS+PkQKBgQCnRZUnLw+JV3EATF/8ZyTmDyQziR/Bk3ALAnJPvIUpdBXla1yH\\nrCOF4G03HXiC+fcYzr21kQOJ4Uo6plLkaiErOkLc66NrLrUXam1uYL/Lv0bPAzLK\\nK0GibHDtl6k+C7V17GbHX17zDLVveWL/6fp4PfrEDqrqgaKk+9PeadYaXwKBgG6m\\nczn0pUVxY60lWrZcCesDeQFMI9rWm8r9fesmGk+teyO8UqBmZswExHZVt5ZgbnKd\\nO1iBDu4YNWJl/l5Y44kVWCC4HaTo8vP1XoQqulGT2sMvZEy1hTBhF6OlwWPh3edJ\\n5bEnHPe3syGZLOET33eR28LtiI6fqB60DSfCKFaRAoGAR74hITKw+PbTdUWql/uT\\nuVHE1JaxhnvNRc+/khoNp903fGAHiVJ5hjnFKRVRUB8uMUtTSfKsS9Y5a4BatvB+\\nAdAY/BHdXad2Xwom8kH9Oirph8exXro3x+FmFzBbwcRwggCGXPX0p1vPPzcZLWnp\\nEXk80T6vA2vVQxYvIrG1yqw=\\n-----END PRIVATE KEY-----\\n\",\n  \"client_email\": \"firebase-adminsdk-i3gy3@lid-rastr.iam.gserviceaccount.com\",\n  \"client_id\": \"105084896569014155165\",\n  \"auth_uri\": \"https://accounts.google.com/o/oauth2/auth\",\n  \"token_uri\": \"https://oauth2.googleapis.com/token\",\n  \"auth_provider_x509_cert_url\": \"https://www.googleapis.com/oauth2/v1/certs\",\n\"client_x509_cert_url\": \"https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-i3gy3%40lid-rastr.iam.gserviceaccount.com\"\n}\n"""

# Pega as configurações do banco do segredo
key_dict = json.loads(textkey_2)
creds = service_account.Credentials.from_service_account_info(key_dict)

# Seleciona o projeto
db = firestore.Client(credentials=creds, project="lid-rastr")

# Ajustando fuso

tz = pytz.timezone('America/Bahia')

##############################################################################
# 												  funcoes					
##############################################################################


def trata_dados(data, tipo):
	# tratamento da planilha de tampas prata
	data.rename(columns={data.columns[0]: "remove"}, inplace=True)
	data.dropna(subset=['remove'], inplace=True)
	data.rename(columns=data.iloc[0].str.strip(), inplace=True)
	data.reset_index(drop=True, inplace=True)
	data.drop([0], inplace=True)
	data.rename(columns={data.columns[17]: "observacao"}, inplace=True)
	data = data.loc[(data['STATUS'].str.lower() == 'armazenada') & (pd.isna(data['observacao']))]
	data = data.iloc[:, [2, 6, 1, 0, 4, 3, 14, 15, 16]]
	
	dicionario_colunas = {
		data.columns[0]: "numero_OT",
		data.columns[1]: "data",
		data.columns[2]: "tipo_bobina",
		data.columns[3]: "codigo_bobina",
		data.columns[4]: "peso_bobina",
		data.columns[5]: "codigo_SAP",
		data.columns[6]: "data_entrada",
		data.columns[7]: "paletes_gerados",
		data.columns[8]: "status"	
	}
	
	data.rename(columns=dicionario_colunas, inplace=True)
	
	data.codigo_SAP = data.codigo_bobina
	
	# define o tipo de tampa de acordo com o parametro tipo
	if tipo == 1:
		data.tipo_bobina = 'Tampa Prata'
	elif tipo == 2:
		data.tipo_bobina = 'Tampa Dourada'
	elif tipo == 3:
		data.tipo_bobina = 'Tampa Branca'
	elif tipo == 4:
		data.tipo_bobina = 'Tampa Lacre Azul'
		
	data.data_entrada = '-'
	data.paletes_gerados = (data['peso_bobina']) * 412 / 187200
	data.paletes_gerados = data.paletes_gerados.astype('int')
	data.status = 'Disponível'
	data['comentario'] = '-'
	data['data_saida'] = '-'
	return data


def upload_excel(uploaded_file):
	# Leitura dos dados do arquivo excel
	try:
		# tratamento da planilha de tampas prata
		df_tp = pd.read_excel(uploaded_file, sheet_name='Bobina Tampa Prata')
		tratado_tp = trata_dados(df_tp, 1)

		# tratamento da planilha de tampass gold
		df_gd = pd.read_excel(uploaded_file, sheet_name='Bobina Tampa Gold')
		tratado_gd = trata_dados(df_gd, 2)

		# tratamento da palnilha de tampas brancas
		df_br = pd.read_excel(uploaded_file, sheet_name='BOBINA TAMPA BRANCA')
		tratado_br = trata_dados(df_br, 3)

		# tratamento da planilha de tampas de lacre azul
		df_ta = pd.read_excel(uploaded_file, sheet_name='Bobina Tampa Lacre Azul')
		tratado_ta = trata_dados(df_ta, 4)

		dados = tratado_tp.append(tratado_gd, ignore_index=True)
		dados = dados.append(tratado_br, ignore_index=True)
		dados = dados.append(tratado_ta, ignore_index=True)
		return dados
	except:
		st.error('Arquivo não compatível')
	return None


def insert_excel(df):
	#try:
	# verifica se há bobinas no sistema
	if df_bobinas.shape[0]:
		# lista de bobinas ja inclusas no sistema
		bobinas_antigas = df_bobinas.numero_OT

		df.numero_OT = df.numero_OT.astype(str)

		# Filtrando os dados (tempo maior que 30 e eventos incluídos em tipo)
		st.subheader('Bobinas a serem inseridas')
		
		df = df[~df['numero_OT'].isin(list(bobinas_antigas))]

	# Se houver variáveis a serem incluídas e faz a inclusão
	if df.shape[0] > 0:
		st.write('Confira os dados antes de inserí-los no sistema. Valores "nan" indicam que faltam dados e a planilha deve ser corrigida.')
		st.write(df)
		batch = db.batch()
		for index, row in df.iterrows():

			# Define a quantidade de paletes que podem ser gerados pela bobina
			qtd_paletes = row.paletes_gerados

			# cria dataframe e preenche com os dados da bobina
			df_paletes_sem = pd.DataFrame(columns=col_pal_sem, index=list(range(qtd_paletes)))
			df_paletes_sem['numero_OT'] = str(row['numero_OT'])
			df_paletes_sem['tipo_tampa'] = str(row['tipo_bobina'])
			df_paletes_sem['data_gerado'] = str(row['data_entrada'])
			df_paletes_sem['data_estoque'] = '-'
			df_paletes_sem['data_consumo'] = '-'
			df_paletes_sem['codigo_SAP'] = '-'
			df_paletes_sem['numero_palete'] = '-'
			df_paletes_sem['codigo_bobina'] = str(row['codigo_SAP'])

			# for para iterar sobre todos os paletes e salvar
			for index, rows in df_paletes_sem.iterrows():
				if index < 10:
						index_str = '0' + str(index)
				else:
						index_str = str(index)
				rows['documento'] = index_str

			row['Paletes'] = df_paletes_sem.to_csv()
			ref = db.collection('Bobina').document(str(row['numero_OT']))
			row_string = row.astype(str)
			batch.set(ref, row_string.to_dict())
		
		inserir = False
		if df.isnull().sum().sum() > 0:
			st.error('Estão faltando dados na planilha, por favor corrigir')
		else:
			inserir = st.button('Inserir os dados no sistema?')
		
		if inserir:
			# escreve os dados no servidor
			batch.commit()	

			# Limpa cache
			caching.clear_cache()		
			return df
		return None
	else:
		st.info('Todas as bobinas filtradas da planilha já estão inseridas no sistema!')
		return None
	#except:
	#	st.error('Dados não inseridos no banco')
	#	return None
	#pass

	
def local_css(file_name):
	with open(file_name) as f:
		st.markdown('<style>{}</style>'.format(f.read()), unsafe_allow_html=True)


local_css("style.css")	


# Define cores para os valores validos ou invalidos
def color(val):
	if val == 'invalido':
		cor = 'red'
	else:
		cor = 'white'
	return 'background-color: %s' % cor


# Gera arquivo excel
def to_excel(df):
	output = BytesIO()
	writer = pd.ExcelWriter(output, engine='xlsxwriter')
	df.to_excel(writer, sheet_name='Sheet1')
	writer.save()
	processed_data = output.getvalue()
	return processed_data


# Gera o link para o download do excel
def get_table_download_link(df):
	"""Generates a link allowing the data in a given panda dataframe to be downloaded
	in:  dataframe
	out: href string
	"""
	val = to_excel(df)
	b64 = base64.b64encode(val)  # val looks like b'...'
	return f'<a href="data:application/octet-stream;base64,{b64.decode()}" download="dados.xlsx">Download dos dados em Excel</a>'  # decode b'abc' => abc


# visualizar pdf
def show_pdf(file_path):
	with open(file_path,"rb") as f:
		base64_pdf = base64.b64encode(f.read()).decode('utf-8')
	pdf_display = f'<embed src="data:application/pdf;base64,{base64_pdf}" width="700" height="1000" type="application/pdf">'
	st.markdown(pdf_display, unsafe_allow_html=True)


def download_etiqueta(data, tipo): # 0 sem selante e 1 com selante

	# carrega arquivo excel base para etiqueta
	wb = load_workbook('teste2.xlsx')

	# seleciona a planilha
	ws = wb.active

	# converte string para datetime
	data['data_estoque'] = pd.to_datetime(data['data_estoque'])

	# sem selante
	if tipo == 0:
		# Preenchimento dos valores
		ws['A7'] = str(data['tipo_tampa'])  # 'tipo produto'
		ws['B7'] = 'Sem selante'  # 'com/sem selante'

		codigo  = str(data['tipo_tampa']) + ' Sem Selante'
		#ws['A9'] = str(data['codigo_bobina'])  # 'codigo produto'
		ws['A9'] = tipos_tampas[codigo]
		ws['B13'] = str(data['numero_OT'])  # numero da bobina
	
	# com selante
	else:
		# Preenchimento dos valores
		ws['A7'] = str(data['tipo_tampa'])  # 'tipo produto'
		ws['B7'] = 'Com selante'  # 'com/sem selante'
		
		codigo  = str(data['tipo_tampa']) + ' Com Selante'
		#ws['A9'] = str(data['codigo_bobina'])  # 'codigo produto'
		ws['A9'] = tipos_tampas[codigo]
		ws['B13'] = str(data['numero_OT'])  # numero da bobina

	# pega a hora que o palete foi para o estoque
	horario = datetime.time(data['data_estoque'])

	# Adequa os valores dos turnos
	if (horario >= time(23, 0, 0)) and (horario < time(7, 0, 0)):
		ws['B11'] = 'A'  # 'turno'
	elif (horario >= time(7, 0, 0)) and (horario < time(15, 0, 0)):
		ws['B11'] = 'B'  # 'turno'
	else:
		ws['B11'] = 'C'  # 'turno'

	ws['A11'] = data['data_estoque']  # 'data'
	ws['C11'] = data['data_estoque']  # 'hora'
	ws['C9'] = data['numero_palete']  # numero etiqueta

	wb.save('teste.xlsx')
	stream = BytesIO()
	wb.save(stream)
	towrite = stream.getvalue()
	b64 = base64.b64encode(towrite).decode()  # some strings

	# link para download e nome do arquivo
	#linko = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="myfilename.xlsx"><span class="highlight blue">Download etiqueta</span></a>'
	linko = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="myfilename.xlsx">Download etiqueta</a>'
	st.markdown(linko, unsafe_allow_html=True)


# leitura de dados do banco
@st.cache(allow_output_mutation=True)
def load_colecoes(colecao, colunas, colunas_pal, tipo):
	# dicionario vazio
	dicionario = {}
	index = 0

	# Define o caminho da coleção do firebase
	posts_ref = db.collection(colecao)

	# Busca todos os documentos presentes na coleção e salva num dataframe
	for doc in posts_ref.stream():
		dic_auxiliar = doc.to_dict()
		dicionario[str(index)] = dic_auxiliar
		if tipo == 1:
			dicionario[str(index)]['documento'] = doc.id
		if tipo == 0:
			dicionario[str(index)]['documento'] = doc.id
		index += 1
	# Transforma o dicionario em dataframe
	df = pd.DataFrame.from_dict(dicionario)

	# troca linhas com colunas
	df = df.T
	df2 = pd.DataFrame(columns=colunas_pal)

	# Bobinas
	if (tipo == 0) and (df.shape[0] > 0):
		# Transforma string em tipo data

		df['data'] = pd.to_datetime(df['data'])

		# Ordena os dados pela data
		df = df.sort_values(by=['data'], ascending=False)

		# Remove o index
		df = df.reset_index(drop=True)

		for index, row in df.iterrows():
			csv = str(row['Paletes'])
			csv_string = StringIO(csv)
			df_aux = pd.read_table(csv_string, sep=',')
			df2 = df2.append(df_aux, ignore_index=True)

		# Ordena as colunas
		df = df[colunas]
		df2 = df2[colunas_pal]
		df2['numero_OT'] = df2['numero_OT'].astype('str')

	# selante
	if (tipo == 1) and (df.shape[0] > 0):
		# Transforma string em tipo data

		df['data'] = pd.to_datetime(df['data'])
		df['data'] = df['data'].dt.strftime('%H:%M %d-%m-%Y')

		# Ordena os dados pela data
		df = df.sort_values(by=['data'], ascending=False)

		# Remove o index
		df = df.reset_index(drop=True)

		for index, row in df.iterrows():
			csv = str(row['Paletes'])
			csv_string = StringIO(csv)
			df_aux = pd.read_table(csv_string, sep=',')
			df2 = df2.append(df_aux, ignore_index=True)

		# Ordena as colunas
		df = df[colunas]
		df2 = df2[colunas_pal]
		df2['numero_lote'] = df2['numero_lote'].astype('str')

	return df, df2

def adicionar_bobina():
	# Formulario para inclusao de bobinas
	dic = {}

	# Dados das bobinas
	with st.form('forms_Bobina'):
		dic['status'] = 'Disponível'
		dic['data'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")
		s1, s2, s3, s4, s6 = st.beta_columns([2.5, 2.5, 2.5, 2.5, 1])
		dic['numero_OT'] = s1.text_input('Número OT')
		dic['tipo_bobina'] = s2.selectbox('Tipo da bobina', list(tipos_bobinas.keys()))
		dic['codigo_bobina'] = tipos_bobinas[dic['tipo_bobina']]
		dic['peso_bobina'] = s3.number_input('Peso da bobina', step=100, format='%i', value=9000, max_value=18000)
		dic['codigo_SAP'] = s4.text_input('Código SAP')
		dic['data_entrada'] = ''
		dic['comentario'] = '-'
		dic['data_saida'] = '-'
		submitted = s6.form_submit_button('Adicionar bobina ao sistema')

	if submitted:
		# verifica se ja existe bobina com o numero de lote inserido
		if df_pal_sem[df_pal_sem['numero_OT'] == (dic['numero_OT'])].shape[0] == 0:
			# Transforma dados do formulário em um dicionário
			keys_values = dic.items()
			new_d = {str(key): str(value) for key, value in keys_values}

			# Verifica campos não preenchidos e os modifica
			for key, value in new_d.items():
				if (value == '') or value == '[]':
					new_d[key] = '-'

			# define a quantidade de paletes gerados pela bobina
			new_d['paletes_gerados'] = int(int(new_d['peso_bobina']) * 412 / 187200)

			# Define a quantidade de paletes que podem ser gerados pela bobina
			qtd_paletes = int(new_d['paletes_gerados'])

			# cria dataframe e preenche com os dados da bobina
			df_paletes_sem = pd.DataFrame(columns=col_pal_sem, index=list(range(qtd_paletes)))
			df_paletes_sem['numero_OT'] = str(new_d['numero_OT'])
			df_paletes_sem['tipo_tampa'] = str(new_d['tipo_bobina'])
			df_paletes_sem['data_gerado'] = str(new_d['data_entrada'])
			df_paletes_sem['data_estoque'] = '-'
			df_paletes_sem['data_consumo'] = '-'
			df_paletes_sem['codigo_SAP'] = '-'
			df_paletes_sem['numero_palete'] = '-'
			df_paletes_sem['codigo_bobina'] =  str(new_d['codigo_SAP'])

			# for para iterar sobre todos os paletes e salvar
			for index, row in df_paletes_sem.iterrows():
				if index < 10:
					index_str = '0' + str(index)
				else:
					index_str = str(index)
				row['documento'] = index_str

			new_d['Paletes'] = df_paletes_sem.to_csv()

			rerun = False
			# Armazena no banco
			try:
				doc_ref = db.collection("Bobina").document(new_d['numero_OT'])
				doc_ref.set(new_d)
				st.success('Bobina adicionada com sucesso!')

				# Limpa cache
				caching.clear_cache()

				# flag para rodar novamente o script
				rerun = True
			except:
				st.error('Falha ao adicionar bobina, tente novamente ou entre em contato com suporte!')

			if rerun:
				st.experimental_rerun()
		else:
			st.error('Já existe bobina com o número do lote informado')

def adicionar_selante():
	# Formulario para inclusao de selante
	dic = {}

	# Dados dos selantes
	with st.form('forms_selante'):
		dic['status'] = 'Disponível'
		dic['data'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")
		s1, s2, s3, s4, s5 = st.beta_columns([2.5, 2.5, 2.5, 2.5, 1])
		dic['numero_lote'] = s1.text_input('Número do lote')
		dic['codigo_SAP'] = s2.text_input('Código SAP')
		dic['peso_vedante'] = s3.number_input('Peso do vedante', step=100, format='%i', value=5000, max_value=10000)
		dic['lote_interno'] = s4.text_input('Lote interno')
		dic['data_entrada'] = '-'
		dic['comentario'] = '-'
		dic['data_saida'] = '-'
		submitted = s5.form_submit_button('Adicionar selante ao sistema')

	if submitted:
		# verifica se ja existe selante com o numero de lote inserido
		if df_pal_com[df_pal_com['numero_lote'] == (dic['numero_lote'])].shape[0] == 0:
			# Transforma dados do formulário em um dicionário
			keys_values = dic.items()
			new_d = {str(key): str(value) for key, value in keys_values}

			# Verifica campos não preenchidos e os modifica
			for key, value in new_d.items():
				if (value == '') or value == '[]':
					new_d[key] = '-'

			# define a quantidade de paletes gerados pelo selante
			new_d['paletes_gerados'] = int(int(new_d['peso_vedante']) * 2857 / 187200)

			# Define a quantidade de paletes que podem ser gerados pelo selante
			qtd_paletes = int(new_d['paletes_gerados'])

			# cria dataframe e preenche com os dados da selante
			df_paletes_selante = pd.DataFrame(columns=col_pal_sel, index=list(range(qtd_paletes)))
			df_paletes_selante['numero_lote'] = str(new_d['numero_lote'])
			df_paletes_selante['codigo_SAP'] = '-'
			df_paletes_selante['data_gerado'] = str(new_d['data_entrada'])
			df_paletes_selante['tipo_tampa'] = '-'
			df_paletes_selante['data_estoque'] = '-'
			df_paletes_selante['data_consumo'] = '-'
			df_paletes_selante['lote_semi'] = '-'
			df_paletes_selante['numero_palete'] = '-'
			df_paletes_selante['codigo_bobina'] = '-'
			df_paletes_selante['numero_OT'] = '-'

			# for para iterar sobre todos os paletes e salvar
			for index, row in df_paletes_selante.iterrows():
				if index < 10:
					index_str = '0' + str(index)
				else:
					index_str = str(index)
				row['documento'] = index_str

			new_d['Paletes'] = df_paletes_selante.to_csv()

			rerun = False
			# Armazena no banco
			try:
				doc_ref = db.collection("Selante").document(new_d['numero_lote'])
				doc_ref.set(new_d)
				st.success('Selante adicionada com sucesso!')

				# Limpa cache
				caching.clear_cache()

				# flag para rodar novamente o script
				rerun = True
			except:
				st.error('Falha ao adicionar selante, tente novamente ou entre em contato com suporte!')

			if rerun:
				st.experimental_rerun()
		else:
			st.error('Já existe selante com o número do lote informado')
###########################################################################################################################################
#####								cofiguracoes aggrid											#######
###########################################################################################################################################
def config_grid(height, df, lim_min, lim_max, customizar):
	sample_size = 12
	grid_height = height

	return_mode = 'AS_INPUT'
	return_mode_value = DataReturnMode.__members__[return_mode]
	# return_mode_value = 'AS_INPUT'

	update_mode = 'VALUE_CHANGED'
	update_mode_value = GridUpdateMode.__members__[update_mode]
	# update_mode_value = 'VALUE_CHANGED'

	# enterprise modules
	enable_enterprise_modules = False
	enable_sidebar = False

	# features
	fit_columns_on_grid_load = customizar
	enable_pagination = False
	paginationAutoSize = False
	use_checkbox = False
	enable_selection = False
	selection_mode = 'single'
	rowMultiSelectWithClick = False
	suppressRowDeselection = False

	if use_checkbox:
		groupSelectsChildren = True
		groupSelectsFiltered = True

	# Infer basic colDefs from dataframe types
	gb = GridOptionsBuilder.from_dataframe(df)

	# customize gridOptions
	if not customizar:
		gb.configure_default_column(groupable=True, value=True, enableRowGroup=True, aggFunc='sum', editable=True)
		gb.configure_column("Medidas", editable=False)
		gb.configure_column('L', editable=False)
		gb.configure_column('V', type=["numericColumn"], precision=5)

		# configures last row to use custom styles based on cell's value, injecting JsCode on components front end
		func_js = """
		function(params) {
			if (params.value > %f) {
			return {
				'color': 'black',
				'backgroundColor': 'orange'
			}
			} else if(params.value <= %f) {
			return {
				'color': 'black',
				'backgroundColor': 'orange'
			}
			} else if((params.value <= %f) && (params.value >= %f)) {
			return {
				'color': 'black',
				'backgroundColor': 'white'
			}
			} else {
			return {
				'color': 'black',
				'backgroundColor': 'red'
			} 
			} 
		};
		""" % (lim_max, lim_min, lim_max, lim_min)

		cellsytle_jscode = JsCode(func_js)

		gb.configure_column('V', cellStyle=cellsytle_jscode)

	if enable_sidebar:
		gb.configure_side_bar()

	if enable_selection:
		gb.configure_selection(selection_mode)
	if use_checkbox:
		gb.configure_selection(selection_mode, use_checkbox=True, groupSelectsChildren=groupSelectsChildren,
							   groupSelectsFiltered=groupSelectsFiltered)
	if ((selection_mode == 'multiple') & (not use_checkbox)):
		gb.configure_selection(selection_mode, use_checkbox=False, rowMultiSelectWithClick=rowMultiSelectWithClick,
							   suppressRowDeselection=suppressRowDeselection)

	if enable_pagination:
		if paginationAutoSize:
			gb.configure_pagination(paginationAutoPageSize=True)
		else:
			gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=paginationPageSize)

	gb.configure_grid_options(domLayout='normal')
	gridOptions = gb.build()
	return gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules


##########################################################################################################
#####								rastreabilidade  		   									  ########
##########################################################################################################

# definicao de colunas para leitura d dados do banco
col_bobinas = ['numero_OT', 'data', 'tipo_bobina', 'codigo_bobina', 'peso_bobina', 'codigo_SAP', 'data_entrada', 'data_saida',
			   'paletes_gerados', 'status', 'comentario']
col_pal_sem = ['numero_OT', 'documento', 'tipo_tampa', 'data_gerado', 'data_estoque', 'data_consumo',
			   'codigo_SAP', 'numero_palete']
col_selante = ['numero_lote', 'lote_interno', 'codigo_SAP', 'peso_vedante', 'data', 'data_entrada', 'data_saida', 'paletes_gerados',
			   'status', 'comentario']
col_pal_sel = ['numero_lote', 'numero_OT', 'documento', 'tipo_tampa', 'codigo_SAP', 'data_gerado', 'data_estoque', 'data_consumo', 'lote_semi', 'numero_palete']

tipos_tampas = {'Tampa Prata Sem Selante': 40009011,
		'Tampa Prata Com Selante': 40009012,
		'Tampa Dourada Sem Selante': 40009013,
		'Tampa Dourada Com Selante': 40009014,
		'Tampa Branca Sem Selante': 40009439,
		'Tampa Branca Com Selante': 40009438}

tipos_bobinas = {'Tampa Prata': 50490760,
		'Tampa Dourada': 50490599,
		'Tampa Branca': 50427252,
		'Tampa Lacre Azul': 50527602}

tipos_selantes = {'Selante': 50491194}

# leitura e exibicao dos dados das bobinas
df_bobinas, df_pal_sem = load_colecoes('Bobina', col_bobinas, col_pal_sem, 0)
df_selantes, df_pal_com = load_colecoes('Selante', col_selante, col_pal_sel, 1)

# define a bobina em uso
if df_bobinas.shape[0] > 0:
	if df_bobinas.loc[df_bobinas['status'] == 'Em uso', 'tipo_bobina'].shape[0] > 0:
		tipo_bobina = df_bobinas.loc[df_bobinas['status'] == 'Em uso', 'tipo_bobina']
		tipo_bobina_uso = str(tipo_bobina.iloc[0])
	else:
		tipo_bobina_uso = 'Não há bobina em uso'
else:
	tipo_bobina_uso = 'Não há bobina em uso'

#verifica se ha bobina em uso
bobina_em_uso = pd.DataFrame()

if df_bobinas.shape[0] > 0:
	bobina_em_uso = df_bobinas[df_bobinas['status'] == 'Em uso']

#verifica se ha selante em uso
selante_em_uso = pd.DataFrame()

if df_selantes.shape[0] > 0:
	selante_em_uso = df_selantes[df_selantes['status'] == 'Em uso']

# dataframes do fifo sem selante
ps_fifo_in = df_pal_sem[(df_pal_sem['data_estoque'] != '-') & (df_pal_sem['data_consumo'] == '-') & (df_pal_sem['tipo_tampa'].astype(str) == tipo_bobina_uso)]
ps_fifo_out = df_pal_sem[df_pal_sem['data_consumo'] != '-']

# dataframes do fifo com selante
sel_fifo_in = df_pal_com[(df_pal_com['data_estoque'] != '-') & (df_pal_com['data_consumo'] == '-') & (df_pal_com['tipo_tampa'].astype(str) == tipo_bobina_uso)] # & (df_pal_com['tipo_tampa'] == str(tipo_bobina_uso))]
sel_fifo_out = df_pal_com[df_pal_com['data_consumo'] != '-']

#######################
# organizacao da tela #
#######################

with st.beta_expander('Gerenciamento de bobinas'):
	st.subheader('Inserir Bobinas')
	uploaded_file = st.file_uploader("Selecione o arquivo Excel para upload")
	if uploaded_file is not None:
		data_excel = upload_excel(uploaded_file)
		df_excel = insert_excel(data_excel)
		df_bobinas = df_bobinas.append(df_excel)

	adicionar_bobina()
	if df_bobinas.shape[0] > 0:
		st.subheader('Selecionar bobina para uso')
		st1, st2 = st.beta_columns([99, 1])

		st.subheader('Detalhamento das bobinas')
		#st.write(df_bobinas)
		gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(198, df_bobinas, 0, 0, True)
		response = AgGrid(
			df_bobinas,
			gridOptions=gridOptions,
			height=grid_height,
			width='100%',
			data_return_mode=return_mode_value,
			update_mode=update_mode_value,
			fit_columns_on_grid_load=fit_columns_on_grid_load,
			allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
			enable_enterprise_modules=enable_enterprise_modules)

with st.beta_expander('Gerenciamento de selantes'):

	st.subheader('Inserir Selante')
	adicionar_selante()

	if df_selantes.shape[0] > 0:
		st.subheader('Selecionar selante para uso')
		st11, st22 = st.beta_columns([99, 1])

		st.subheader('Detalhamento dos selantes')

		gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(199, df_selantes, 0, 0, True)
		response = AgGrid(
			df_selantes,
			gridOptions=gridOptions,
			height=grid_height,
			width='100%',
			data_return_mode=return_mode_value,
			update_mode=update_mode_value,
			fit_columns_on_grid_load=fit_columns_on_grid_load,
			allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
			enable_enterprise_modules=enable_enterprise_modules)

# define imagem e barra lateral
col2, imagem, col4 = st.beta_columns([3, 10, 3])
imagem.markdown("<h1 style='text-align: center; color: gray;'>Tipo de tampa em produção: {}</h1>".format(tipo_bobina_uso), unsafe_allow_html=True)
imagem.image('lid_linha.png')

recursos = ['Remover bobinas ou selantes', 
'Histórico de paletes sem selante',
'Histórico de paletes com selante',
'Apontamento de código SAP',
'Detalhamento de bobinas e selantes por data'
]
telas = imagem.radio('Selecione o recurso que deseja utilizar', recursos)

# verifica se há bobinas no sistema para habilitar as demais funcionalidades do sistema
if df_bobinas.shape[0] > 0:

	###########################################
	# Selecionar bobinas disponiveis para uso #
	###########################################

	# Verifica bobinas disponiveis
	df_bobinas_disp = df_bobinas[df_bobinas['status'] == 'Disponível']
	df_bobinas_disp['data'] = pd.to_datetime(df_bobinas_disp['data'])
	df_bobinas_disp.sort_values(by=['data'], inplace=True)

	# cria selectbox para selecionar bobinas
	numero_bobina_full = st1.selectbox('Selecione a próxima bobina:', list((df_bobinas_disp['numero_OT'].astype(str) + ' / Data: ' + df_bobinas_disp['data'].dt.strftime("%d/%m/%Y") + ' / Tipo: ' + df_bobinas_disp['tipo_bobina'].astype(str))))
	numero_bobina = numero_bobina_full.split()[0]

	# parte do principio que nenhuma bobina foi selecionada
	selecionar_bobina = False

	# verifica se foi selecionada alguma bobina
	if numero_bobina != None:
		selecionar_bobina = st1.button('Utilizar a bobina selecionada?')
	else:
		st1.info('Não há bobinas disponiveis')

	if selecionar_bobina:

		###################################
		# Coloca anterior como finalizada #
		###################################

		if bobina_em_uso.shape[0] > 0:

			# seleciona a bobina em uso
			val_em_uso = bobina_em_uso.iloc[0,0]

			# modifica bobina selecionada para finalizada
			df_bobinas.loc[df_bobinas['numero_OT'] == val_em_uso, 'status'] = 'Finalizada'
			df_bobinas.loc[df_bobinas['numero_OT'] == val_em_uso, 'data_saida'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")

			# prepara dados para escrever no banco
			dic_fin = {}
			dic_fin = df_bobinas.loc[df_bobinas['numero_OT'] == val_em_uso].to_dict('records')

			# Transforma dados do formulário em um dicionário
			keys_values = dic_fin[0].items()
			new_fin = {str(key): str(value) for key, value in keys_values}
			documento = new_fin['numero_OT']

			# escreve o dataframe dos paletes na bobina para escrita em banco (não altera valor, mas escreve para não perder os dados)
			new_fin['Paletes'] = df_pal_sem[df_pal_sem['numero_OT'] == val_em_uso].to_csv()

			# Armazena no banco as alteracoes na bobina
			try:
				doc_ref = db.collection("Bobina").document(documento)
				doc_ref.set(new_fin)
				st.success('Formulário armazenado com sucesso!')
			except:
				st.error('Falha ao armazenar formulário, tente novamente ou entre em contato com suporte!')
				caching.clear_cache()
		else:
			st.info('Não havia bobina em uso!')

		####################################
		# Coloca bobina selecionada em uso #
		####################################

		st.write(numero_bobina)
		st.write(df_bobinas['numero_OT'])

		# modifica bobina selecionada para uso
		df_bobinas.loc[df_bobinas['numero_OT'] == numero_bobina, 'status'] = 'Em uso'
		df_bobinas.loc[df_bobinas['numero_OT'] == numero_bobina, 'data_entrada'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")

		# prepara dados para escrever no banco
		dic_bobina_uso = {}
		dic_bobina_uso = df_bobinas.loc[df_bobinas['numero_OT'] == numero_bobina].to_dict('records')

		# Transforma dados do formulário em um dicionário
		keys_values = dic_bobina_uso[0].items()
		new_uso = {str(key): str(value) for key, value in keys_values}
		documento = new_uso['numero_OT']

		# Filtra paletes da bobina em uso e atualiza valores
		df_pal_sem.loc[df_pal_sem['numero_OT'] == numero_bobina, 'data_gerado'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")

		st.write(df_pal_sem['numero_palete'])

		# # Coloca o numero dos paletes
		# if (df_pal_sem['numero_palete'] != '-').any():
		# 	maximo_index = int(df_pal_sem.loc[df_pal_sem['numero_palete'] != '-', 'numero_palete'].max()) + 1
		# 	df_pal_sem.loc[df_pal_sem['numero_OT'] == numero_bobina, 'numero_palete'] = df_pal_sem['documento'] + maximo_index
		# else:
		# 	# modificar o valor no final para adequar a realidade da linha rodando
		# 	df_pal_sem.loc[df_pal_sem['numero_OT'] == numero_bobina, 'numero_palete'] = df_pal_sem['documento'] + 1000 

		# Escreve o dataframe dos paletes na bobina para escrita em banco
		new_uso['Paletes'] = df_pal_sem[df_pal_sem['numero_OT'] == numero_bobina].to_csv()

		# Flag de rerun da aplicacao
		flag_rerun = False

		# Armazena no banco as alteracoes na bobina
		try:
			doc_ref = db.collection("Bobina").document(documento)
			doc_ref.set(new_uso)
			st.success('Formulário armazenado com sucesso!')
			flag_rerun = True
		except:
			st.error('Falha ao armazenar formulário, tente novamente ou entre em contato com suporte!')
			caching.clear_cache()

		if flag_rerun:
			st.experimental_rerun()

	############################
	# FIFO paletes sem selante #
	############################

	# Adiciona paletes
	with col2:
		st.subheader('Sem selante')

		if (ps_fifo_in.shape[0] < 5) & (df_bobinas[df_bobinas['status'] == 'Em uso'].shape[0] > 0):
			add_palete_sem = col2.button('Gerar palete TP sem Selante')
			if add_palete_sem:

				# identifica o ultimo numero de palete utilizado
				maximo_index_s = 1000
				if (df_pal_sem['numero_palete'] != '-').any():
					maximo_index_aux = df_pal_sem.loc[df_pal_sem['numero_palete'] != '-', 'numero_palete']
					maximo_index_s = int(maximo_index_aux.astype('int').max()) + 1

				# atribuir numero ao palete
				bobina_atual = df_bobinas[df_bobinas['status'] == 'Em uso']['numero_OT']
				df_temp = df_pal_sem.loc[(df_pal_sem['numero_OT'] == bobina_atual.iloc[0]) & (df_pal_sem['data_estoque'] == '-') & (df_pal_sem['numero_palete'] == '-')]
				df_temp.iloc[0, 7] = maximo_index_s
				df_pal_sem.loc[(df_pal_sem['numero_OT'] == bobina_atual.iloc[0]) & (df_pal_sem['data_estoque'] == '-') & (df_pal_sem['numero_palete'] == '-')] = df_temp

				# verificar selante em uso
				numero_palete = maximo_index_s

				# atualiza data de estoque do palete
				df_pal_sem.loc[df_pal_sem['numero_palete'] == numero_palete, 'data_estoque'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")

				# prepara dados para escrever no banco
				dic_fifo_in = {}
				dic_fifo_in = df_bobinas.loc[df_bobinas['numero_OT'] == bobina_atual.iloc[0]].to_dict('records')

				# Transforma dados do formulário em um dicionário
				keys_values = dic_fifo_in[0].items()
				new_fifo_in = {str(key): str(value) for key, value in keys_values}
				documento = new_fifo_in['numero_OT']

				# Escreve o dataframe dos paletes na bobina para escrita em banco
				new_fifo_in['Paletes'] = df_pal_sem[df_pal_sem['numero_OT'] == bobina_atual.iloc[0]].to_csv()

				# Flag de rerun da aplicacao
				flag_rerun = False

				# Armazena no banco as alteracoes na bobina
				try:
					doc_ref = db.collection("Bobina").document(documento)
					doc_ref.set(new_fifo_in)
					flag_rerun = True

				except:
					st.error('Falha ao atualizar informacoes do palete, tente novamente ou entre em contato com suporte!')

				if flag_rerun:
					st.experimental_rerun()

		elif (ps_fifo_in.shape[0] >= 5):
			st.error('Há paletes demais na reserva')
		
		fifo_in_show = ps_fifo_in.sort_values(by='data_estoque', ascending=True)[['numero_palete', 'tipo_tampa']]
		fifo_in_show.rename(columns={'numero_palete': 'Gerados'}, inplace=True)

		if fifo_in_show.shape[0] > 0:
			gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(175, fifo_in_show, 0, 0, True)
			response = AgGrid(
				fifo_in_show,
				gridOptions=gridOptions,
				height=grid_height,
				width='100%',
				data_return_mode=return_mode_value,
				update_mode=update_mode_value,
				fit_columns_on_grid_load=fit_columns_on_grid_load,
				allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
				enable_enterprise_modules=enable_enterprise_modules)
			st.info(':exclamation: **Próximo palete: ' + str(fifo_in_show.iloc[0, 0]) + '**')

		# consome paletes
		if ps_fifo_in.shape[0] > 0:
			# download da etiqueta
			download_etiqueta(ps_fifo_in.sort_values(by='data_estoque', ascending=True).iloc[0], 0)

			con_palete_sem = col2.button('Consumir palete TP sem Selante')
			if con_palete_sem:
				# observa o indice do primeiro elemento do fifo
				numero_palete_aux = ps_fifo_in.sort_values(by='data_estoque', ascending=True).iloc[0]
				numero_palete = numero_palete_aux.iloc[7]

				# atualiza a data de consumo do palete consumido
				df_pal_sem.loc[(df_pal_sem['numero_palete'] == numero_palete), 'data_consumo'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")

				#identifica o numero da bobina do palete
				bobina_consumo = df_pal_sem.loc[(df_pal_sem['numero_palete'] == numero_palete), 'numero_OT']

				# prepara dados para escrever no banco
				dic_fifo_out = {}
				dic_fifo_out = df_bobinas.loc[df_bobinas['numero_OT'] == bobina_consumo.iloc[0]].to_dict('records')

				# Transforma dados do formulário em um dicionário
				keys_values = dic_fifo_out[0].items()
				new_fifo_out = {str(key): str(value) for key, value in keys_values}
				documento = new_fifo_out['numero_OT']

				# Escreve o dataframe dos paletes na bobina para escrita em banco
				new_fifo_out['Paletes'] = df_pal_sem[df_pal_sem['numero_OT'] == bobina_consumo.iloc[0]].to_csv()

				# Flag de rerun da aplicacao
				flag_rerun = False

				# Armazena no banco as alteracoes na bobina
				try:
					doc_ref = db.collection("Bobina").document(documento)
					doc_ref.set(new_fifo_out)
					flag_rerun = True

				except:
					st.error('Falha ao atualizar informacoes do palete, tente novamente ou entre em contato com suporte!')

				if flag_rerun:
					st.experimental_rerun()

		else:
			st.error('Não há palete sem selante para consumir')

		fifo_out_show = ps_fifo_out.sort_values(by='data_consumo', ascending=False)[['numero_palete', 'tipo_tampa']]
		fifo_out_show.rename(columns={'numero_palete': 'Consumidos'}, inplace=True)
		
		if fifo_out_show.shape[0] > 0:
			gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(175, fifo_out_show, 0, 0, True)
			response = AgGrid(
				fifo_out_show,
				gridOptions=gridOptions,
				height=grid_height,
				width='100%',
				data_return_mode=return_mode_value,
				update_mode=update_mode_value,
				fit_columns_on_grid_load=fit_columns_on_grid_load,
				allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
				enable_enterprise_modules=enable_enterprise_modules)
		else:
			st.info('Não foram consumidos paletes sem selante')

	###########################################
	# Selecionar selantes disponiveis para uso#
	###########################################
	if df_selantes[df_selantes['status'] == 'Disponível'].shape[0] > 0:
		# Verifica selantes disponiveis
		df_selantes_disp = df_selantes[df_selantes['status'] == 'Disponível']
		df_selantes_disp['data'] = pd.to_datetime(df_selantes_disp['data'])
		df_selantes_disp.sort_values(by=['data'], inplace=True)

		# cria selectbox para selecionar selantes
		numero_selante_full = st11.selectbox('Selecione o próximo selante:', list((df_selantes_disp['numero_lote'].astype(str) + ' / Data: ' + df_selantes_disp['data'].dt.strftime("%d/%m/%Y"))))
		numero_selante = numero_selante_full.split()[0]
		#numero_selante = st11.selectbox('Selecione o próximo selante:', list(df_selantes_disp['numero_lote']))

		# parte do principio que nenhuma selante foi selecionada
		selecionar_selante = False

		# verifica se foi selecionada alguma selante
		if numero_selante != None:
			selecionar_selante = st11.button('Utilizar o selante selecionado?')
		else:
			st11.info('Não há selantes disponiveis')

		if selecionar_selante:

			###################################
			# Coloca anterior como finalizada #
			###################################

			if selante_em_uso.shape[0] > 0:

				# seleciona a selante em uso
				val_em_uso = selante_em_uso.iloc[0,0]

				# modifica selante selecionada para finalizada
				df_selantes.loc[df_selantes['numero_lote'] == val_em_uso, 'status'] = 'Finalizada'
				df_selantes.loc[df_selantes['numero_lote'] == val_em_uso, 'data_entrada'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")
				df_selantes.loc[df_selantes['numero_lote'] == val_em_uso, 'data_saida'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")

				# prepara dados para escrever no banco
				dic_fin = {}
				dic_fin = df_selantes.loc[df_selantes['numero_lote'] == val_em_uso].to_dict('records')

				# Transforma dados do formulário em um dicionário
				keys_values = dic_fin[0].items()
				new_fin = {str(key): str(value) for key, value in keys_values}
				documento = new_fin['numero_lote']

				# escreve o dataframe dos paletes na selante para escrita em banco (não altera valor, mas escreve para não perder os dados)
				new_fin['Paletes'] = df_pal_com[df_pal_com['numero_lote'] == val_em_uso].to_csv()

				# Armazena no banco as alteracoes na selante
				try:
					doc_ref = db.collection("Selante").document(documento)
					doc_ref.set(new_fin)
					st.success('Formulário armazenado com sucesso!')
				except:
					st.error('Falha ao armazenar formulário, tente novamente ou entre em contato com suporte!')
					caching.clear_cache()
			else:
				st.info('Não havia selante em uso!')

			#####################################
			# Coloca selante selecionada em uso #
			#####################################

			# modifica selante selecionada para uso
			df_selantes.loc[df_selantes['numero_lote'] == numero_selante, 'status'] = 'Em uso'
			df_selantes.loc[df_selantes['numero_lote'] == numero_selante, 'data_entrada'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")

			# prepara dados para escrever no banco
			dic_selante_uso = {}
			dic_selante_uso = df_selantes.loc[df_selantes['numero_lote'] == numero_selante].to_dict('records')

			# Transforma dados do formulário em um dicionário
			keys_values = dic_selante_uso[0].items()
			new_uso = {str(key): str(value) for key, value in keys_values}
			documento = new_uso['numero_lote']

			# Filtra paletes da selante em uso e atualiza valores
			df_pal_com.loc[df_pal_com['numero_lote'] == numero_selante, 'data_gerado'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")

			# # Coloca o numero dos paletes
			# if (df_pal_com['numero_palete'] != '-').any():
			# 	maximo_index = int(df_pal_com.loc[df_pal_com['numero_palete'] != '-', 'numero_palete'].max()) + 1
			# 	df_pal_com.loc[df_pal_com['numero_lote'] == numero_selante, 'numero_palete'] = df_pal_com['documento'] + maximo_index
			# else:
			# 	df_pal_com.loc[df_pal_com['numero_lote'] == numero_selante, 'numero_palete'] = df_pal_com['documento']

			# Escreve o dataframe dos paletes na selante para escrita em banco
			new_uso['Paletes'] = df_pal_com[df_pal_com['numero_lote'] == numero_selante].to_csv()

			# Flag de rerun da aplicacao
			flag_rerun = False

			# Armazena no banco as alteracoes na selante
			try:
				doc_ref = db.collection("Selante").document(documento)
				doc_ref.set(new_uso)
				st.success('Formulário armazenado com sucesso!')
				flag_rerun = True
			except:
				st.error('Falha ao armazenar formulário, tente novamente ou entre em contato com suporte!')
				caching.clear_cache()

			if flag_rerun:
				st.experimental_rerun()

		##############################
		# fifo_s paletes com selante #
		##############################

		# Adiciona paletes
		with col4:
			st.subheader('Com selante')

			if (sel_fifo_in.shape[0] < 5) & (df_selantes[df_selantes['status'] == 'Em uso'].shape[0] > 0):
				add_palete_sem = col4.button('Gerar palete TP com Selante')
				if add_palete_sem:

					# identifica o ultimo numero de palete utilizado
					maximo_index = 0
					if (df_pal_com['numero_palete'] != '-').any():
						maximo_index_aux = df_pal_com.loc[df_pal_com['numero_palete'] != '-', 'numero_palete']
						maximo_index = int(maximo_index_aux.astype('int').max()) + 1

					# atribuir numero ao palete
					selante_atual = df_selantes[df_selantes['status'] == 'Em uso']['numero_lote']
					df_temp = df_pal_com.loc[(df_pal_com['numero_lote'] == selante_atual.iloc[0]) & (df_pal_com['data_estoque'] == '-') & (df_pal_com['numero_palete'] == '-')]
					df_temp.iloc[0, 9] = maximo_index
					df_pal_com.loc[(df_pal_com['numero_lote'] == selante_atual.iloc[0]) & (df_pal_com['data_estoque'] == '-') & (df_pal_com['numero_palete'] == '-')] = df_temp

					# verificar selante em uso
					numero_palete = maximo_index

					# atualiza valores de data de estoque e o tipo de tampa
					df_pal_com.loc[df_pal_com['numero_palete'] == numero_palete, 'data_estoque'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")
					df_pal_com.loc[df_pal_com['numero_palete'] == numero_palete, 'tipo_tampa'] = tipo_bobina_uso
					df_pal_com.loc[df_pal_com['numero_palete'] == numero_palete, 'numero_OT'] = bobina_em_uso.iloc[0,0]

					# prepara dados para escrever no banco
					dic_fifo_s_in = {}
					dic_fifo_s_in = df_selantes.loc[df_selantes['numero_lote'] == selante_atual.iloc[0]].to_dict('records')

					# Transforma dados do formulário em um dicionário
					keys_values = dic_fifo_s_in[0].items()
					new_fifo_s_in = {str(key): str(value) for key, value in keys_values}
					documento = new_fifo_s_in['numero_lote']

					# Escreve o dataframe dos paletes na selante para escrita em banco
					new_fifo_s_in['Paletes'] = df_pal_com[df_pal_com['numero_lote'] == selante_atual.iloc[0]].to_csv()

					# Flag de rerun da aplicacao
					flag_rerun = False

					# Armazena no banco as alteracoes na selante
					try:
						doc_ref = db.collection("Selante").document(documento)
						doc_ref.set(new_fifo_s_in)
						flag_rerun = True

					except:
						st.error('Falha ao atualizar informacoes do palete, tente novamente ou entre em contato com suporte!')

					if flag_rerun:
						st.experimental_rerun()

			elif (sel_fifo_in.shape[0] >= 5):
				st.error('há paletes demais na reserva')

			fifo_s_in_show = sel_fifo_in.sort_values(by='data_estoque', ascending=True)[['numero_palete', 'tipo_tampa']]
			fifo_s_in_show.rename(columns={'numero_palete': 'Gerados'}, inplace=True)

			if fifo_s_in_show.shape[0] > 0:
				gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(175, fifo_s_in_show, 0, 0, True)
				response = AgGrid(
					fifo_s_in_show,
					gridOptions=gridOptions,
					height=grid_height,
					width='100%',
					data_return_mode=return_mode_value,
					update_mode=update_mode_value,
					fit_columns_on_grid_load=fit_columns_on_grid_load,
					allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
					enable_enterprise_modules=enable_enterprise_modules)
				st.success(':exclamation: **Próximo palete: ' + str(fifo_s_in_show.iloc[0, 0]) + '**')

			# consome paletes
			if sel_fifo_in.shape[0] > 0:
				# download das etiquetas
				download_etiqueta(sel_fifo_in.sort_values(by='data_estoque', ascending=True).iloc[0], 1)

				con_palete_sem = col4.button('Consumir palete TP com Selante')
				if con_palete_sem:
					# observa o indice do primeiro elemento do fifo_s
					numero_palete_aux = sel_fifo_in.sort_values(by='data_estoque', ascending=True).iloc[0]
					numero_palete = numero_palete_aux.iloc[9]

					# atualiza a data de consumo do palete consumido
					df_pal_com.loc[(df_pal_com['numero_palete'] == numero_palete), 'data_consumo'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")

					#identifica o numero da selante do palete
					selante_consumo = df_pal_com.loc[(df_pal_com['numero_palete'] == numero_palete), 'numero_lote']

					# prepara dados para escrever no banco
					dic_fifo_s_out = {}
					dic_fifo_s_out = df_selantes.loc[df_selantes['numero_lote'] == selante_consumo.iloc[0]].to_dict('records')

					# Transforma dados do formulário em um dicionário
					keys_values = dic_fifo_s_out[0].items()
					new_fifo_s_out = {str(key): str(value) for key, value in keys_values}
					documento = new_fifo_s_out['numero_lote']

					# Escreve o dataframe dos paletes na selante para escrita em banco
					new_fifo_s_out['Paletes'] = df_pal_com[df_pal_com['numero_lote'] == selante_consumo.iloc[0]].to_csv()

					# Flag de rerun da aplicacao
					flag_rerun = False

					# Armazena no banco as alteracoes na selante
					try:
						doc_ref = db.collection("Selante").document(documento)
						doc_ref.set(new_fifo_s_out)
						flag_rerun = True

					except:
						st.error('Falha ao atualizar informações do palete, tente novamente ou entre em contato com suporte!')

					if flag_rerun:
						st.experimental_rerun()

			else:
				st.error('Não há palete com selante para consumir')

			fifo_s_out_show = sel_fifo_out.sort_values(by='data_consumo', ascending=False)[['numero_palete', 'tipo_tampa']]
			fifo_s_out_show.rename(columns={'numero_palete': 'Consumidos'}, inplace=True)

			if fifo_s_out_show.shape[0] > 0:
				gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(175, fifo_s_out_show, 0, 0, True)
				response = AgGrid(
					fifo_s_out_show,
					gridOptions=gridOptions,
					height=grid_height,
					width='100%',
					data_return_mode=return_mode_value,
					update_mode=update_mode_value,
					fit_columns_on_grid_load=fit_columns_on_grid_load,
					allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
					enable_enterprise_modules=enable_enterprise_modules)
			else:
				st.info('Não foram consumidos paletes com selante')

if telas == 'Remover bobinas ou selantes':
	st.subheader('Remoção de bobinas e selantes da produção')
	# colunas para remoção de bobinas e colunas
	t0, space1, t1 = st.beta_columns([12, 0.5, 12])
	c0, c1, c2, space2, c3, c4, c5 = st.beta_columns([3.5,1.5,1,0.5,3.5,1.5,1])

	# titulos
	t0.subheader('Remover bobinas')
	t1.subheader('Remover selante')

	if bobina_em_uso.shape[0] > 0:
		# coleta os dados relativos a remoção da bobina
		comentario_remover = c0.text_input('Descreva o motivo da retirada da bobina')
		peso_remover = c1.number_input('Peso restante', format='%i', value=5000, step=100)
		remover_bobina = c2.button('Remover bobina em uso')

		if remover_bobina:
			# concatena comentario e peso para escrita no banco
			comentario_peso = ('Motivo: ' + comentario_remover + ' Peso restante: ' + str(peso_remover))

			# seleciona a bobina em uso
			val_em_uso = bobina_em_uso.iloc[0,0]

			# modifica bobina selecionada para removida
			df_bobinas.loc[df_bobinas['numero_OT'] == val_em_uso, 'status'] = 'Removida'
			df_bobinas.loc[df_bobinas['numero_OT'] == val_em_uso, 'data_saida'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")
			df_bobinas.loc[df_bobinas['numero_OT'] == val_em_uso, 'comentario'] = comentario_peso

			# peso incial da bobinaa
			peso_inicial = bobina_em_uso.iloc[0,4]

			# calculo do peso consumido
			peso_consumido = int(peso_inicial) - peso_remover

			# paletes produzidos no total antes da remoção
			paletes_produzidos = int((peso_consumido) * 412 / 187200)

			# atualiza o total de paletes produzidos pela bobina
			df_bobinas.loc[df_bobinas['numero_OT'] == val_em_uso, 'paletes_gerados'] = paletes_produzidos

			# remove da lista da paletes os paletes que não foram gerados
			df_pal_sem.drop(df_pal_sem.loc[(df_pal_sem['numero_OT'] == val_em_uso) & (df_pal_sem['documento'] >= paletes_produzidos)].index, inplace=True)

			# prepara dados para escrever no banco
			dic_remove = {}
			dic_remove = df_bobinas.loc[(df_bobinas['numero_OT'] == val_em_uso)].to_dict('records')

			# Transforma dados do formulário em um dicionário
			keys_values = dic_remove[0].items()
			new_remove = {str(key): str(value) for key, value in keys_values}
			documento_remove = new_remove['numero_OT']

			# escreve o dataframe dos paletes na selante para escrita em banco (não altera valor, mas escreve para não perder os dados)
			new_remove['Paletes'] = df_pal_sem.loc[(df_pal_sem['numero_OT'] == val_em_uso)].to_csv()

			# flag para rodar novamente o script
			rerun = False

			# Armazena no banco as alteracoes da bobina
			try:
				doc_ref = db.collection("Bobina").document(documento_remove)
				doc_ref.set(new_remove)
				st.success('Modificação armazenada com sucesso!')
				rerun = True
			except:
				st.error('Falha ao armazenar modificação, tente novamente ou entre em contato com suporte!')
				caching.clear_cache()

			# comando para rodar novament o script
			if rerun:
				st.experimental_rerun()

	else:
		c0.info('Não há bobina em uso')

	if selante_em_uso.shape[0] > 0:
		# coleta os dados relativos a remoção do selante
		comentario_remover_sel = c3.text_input('Descreva o motivo da retirada do selante')
		peso_remover_sel = c4.number_input('Peso restante', format='%i', value=3000, step=100)
		remover_selante = c5.button('Remover selante em uso')

		if remover_selante:
			# concatena comentario e peso para escrita no banco
			comentario_peso_sel = ('Motivo: ' + comentario_remover_sel + ' Peso restante: ' + str(peso_remover_sel))

			# verificar selante em uso
			selante_atual = selante_em_uso.iloc[0,0]

			# modifica bobina selecionada para removida
			df_selantes.loc[df_selantes['numero_lote'] == selante_atual, 'status'] = 'Removida'
			df_selantes.loc[df_selantes['numero_lote'] == selante_atual, 'data_saida'] = datetime.now(tz).strftime("%H:%M %d-%m-%Y")
			df_selantes.loc[df_selantes['numero_lote'] == selante_atual, 'comentario'] = comentario_peso_sel

			# peso incial da bobinaa
			peso_inicial_sel = selante_em_uso.iloc[0,3]

			# calculo do peso consumido
			peso_consumido_sel = int(peso_inicial_sel) - peso_remover_sel

			# paletes produzidos no total antes da remoção
			paletes_produzidos_sel = int((peso_consumido_sel) * 2857 / 187200)

			# atualiza o total de paletes produzidos pela bobina
			df_selantes.loc[df_selantes['numero_lote'] == selante_atual, 'paletes_gerados'] = paletes_produzidos_sel

			# remove da lista da paletes os paletes que não foram gerados
			df_pal_com.drop(df_pal_com.loc[(df_pal_com['numero_lote'] == selante_atual) & (df_pal_com['documento'] >= paletes_produzidos_sel)].index, inplace=True)

			# prepara dados para escrever no banco
			dic_remove = {}
			dic_remove = df_selantes.loc[(df_selantes['numero_lote'] == selante_atual)].to_dict('records')

			# Transforma dados do formulário em um dicionário
			keys_values = dic_remove[0].items()
			new_remove_sel = {str(key): str(value) for key, value in keys_values}
			documento_remove_sel = new_remove_sel['numero_lote']

			# escreve o dataframe dos paletes na selante para escrita em banco (não altera valor, mas escreve para não perder os dados)
			new_remove_sel['Paletes'] = df_pal_com.loc[(df_pal_com['numero_lote'] == selante_atual)].to_csv()

			# flag para rodar novamente o script
			rerun = False

			# Armazena no banco as alteracoes da selante
			try:
				doc_ref = db.collection("Selante").document(documento_remove_sel)
				doc_ref.set(new_remove_sel)
				st.success('Modificação armazenada com sucesso!')
				rerun = True
			except:
				st.error('Falha ao armazenar modificação, tente novamente ou entre em contato com suporte!')
				caching.clear_cache()
			
			# comando para rodar novament o script
			if rerun:
				st.experimental_rerun()
	else:
		c3.info('Não há selante em uso')

if telas == 'Histórico de paletes sem selante':
	if df_bobinas.shape[0] > 0:
		st.subheader('Histórico de paletes sem selante')
		gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(400, df_pal_sem, 0, 0, True)
		response = AgGrid(
			df_pal_sem,
			gridOptions=gridOptions,
			height=grid_height,
			width='100%',
			data_return_mode=return_mode_value,
			update_mode=update_mode_value,
			fit_columns_on_grid_load=fit_columns_on_grid_load,
			allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
			enable_enterprise_modules=enable_enterprise_modules)

if telas == 'Histórico de paletes com selante':
	if df_selantes.shape[0] > 0:
			st.subheader('Histórico de paletes com selante')
			gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(400, df_pal_com, 0, 0, True)
			response = AgGrid(
				df_pal_com,
				gridOptions=gridOptions,
				height=grid_height,
				width='100%',
				data_return_mode=return_mode_value,
				update_mode=update_mode_value,
				fit_columns_on_grid_load=fit_columns_on_grid_load,
				allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
				enable_enterprise_modules=enable_enterprise_modules)

if telas == 'Detalhamento de bobinas e selantes por data':
	st.subheader('Detalhamento de bobinas por data')

	data_filtro = st.date_input('Selecione a data que deseja filtrar')
	st.subheader('Bobinas utilizadas na data selecionada')
	if df_bobinas.shape[0] > 0:
		# bobinas que possuem data de entrada e de saída
		bobinas_filtradas = df_bobinas.loc[(df_bobinas['data_entrada'] != '-') ] 
		bobinas_filtradas_s = df_bobinas.loc[df_bobinas['data_saida'] != '-']

		if (bobinas_filtradas.shape[0] > 0) or (bobinas_filtradas_s.shape[0] > 0):
			# converte os valores de string para datetime
			bobinas_filtradas['data_entrada'] = pd.to_datetime(bobinas_filtradas['data_entrada'])
			bobinas_filtradas_s['data_saida'] = pd.to_datetime(bobinas_filtradas_s['data_saida'])

			# filtra as bobinas de acordo com a data
			filtro_bobina = bobinas_filtradas.loc[(bobinas_filtradas['data_entrada'].dt.date == data_filtro)]
			filtro_bobina_s = bobinas_filtradas_s.loc[bobinas_filtradas_s['data_saida'].dt.date == data_filtro]
			
			if (filtro_bobina.shape[0] > 0) or (filtro_bobina_s.shape[0] > 0):
				
				# combina os dados de entrada e os de saida
				resultado = filtro_bobina.append(filtro_bobina_s)
				
				# transforma as datas de volta em strings para facilitar a visualizacao
				resultado['data_entrada'] = pd.to_datetime(resultado['data_entrada'])
				resultado['data'] = pd.to_datetime(resultado['data'])

				# ordena os valores pela data de entrada
				resultado = resultado.sort_values(by='data_entrada')

				# remove os dados duplicados de acordo com a coluna numero_ot
				resultado = resultado.drop_duplicates(subset='numero_OT')

				# organiza os dados para exibição
				#resultado['data_saida'] = resultado['data_saida'].apply(lambda x: '-' if x == '-' else x.dt.strftime("%H:%M %d/%m/%Y"))
				resultado['data'] = resultado['data'].dt.strftime("%d/%m/%Y")
				resultado['data_entrada'] = resultado['data_entrada'].dt.strftime("%H:%M %d/%m/%Y")

				#st.table(resultado)
				gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(120, resultado, 0, 0, True)
				response = AgGrid(
					resultado,
					gridOptions=gridOptions,
					height=grid_height,
					width='100%',
					data_return_mode=return_mode_value,
					update_mode=update_mode_value,
					fit_columns_on_grid_load=fit_columns_on_grid_load,
					allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
					enable_enterprise_modules=enable_enterprise_modules)
			else:
				st.error('Não há bobinas utilizadas na data selecionada')
		else:
			st.error('Não há bobinas utilizadas na data selecionada')
	else:
		st.error('Não há bobinas utilizadas na data selecionada')

	st.subheader('Paletes sem selante utilizados na data selecionada')
	if df_pal_sem.shape[0] > 0:
		# bobinas que possuem data de entrada e de saída
		p_sem_filtrado = df_pal_sem.loc[df_pal_sem['data_consumo'] != '-']

		if p_sem_filtrado.shape[0] > 0:
			# converte os valores de string para datetime
			p_sem_filtrado['data_consumo'] = pd.to_datetime(p_sem_filtrado['data_consumo'])
			p_sem_filtrado['data_gerado'] = pd.to_datetime(p_sem_filtrado['data_gerado'])
			p_sem_filtrado['data_estoque'] = pd.to_datetime(p_sem_filtrado['data_estoque'])

			# filtra as bobinas de acordo com a data
			filtro_pal_sem = p_sem_filtrado.loc[(p_sem_filtrado['data_consumo'].dt.date == data_filtro)]

			if filtro_pal_sem.shape[0] > 0:
				# transforma as datas de volta em strings para facilitar a visualizacao
				filtro_pal_sem['data_gerado'] = filtro_pal_sem['data_gerado'].dt.strftime("%d/%m/%Y")
				filtro_pal_sem['data_estoque'] = filtro_pal_sem['data_estoque'].dt.strftime("%H:%M %d/%m/%Y")
				filtro_pal_sem['data_consumo'] = filtro_pal_sem['data_consumo'].dt.strftime("%H:%M %d/%m/%Y")

				gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(120, filtro_pal_sem, 0, 0, True)
				response = AgGrid(
					filtro_pal_sem,
					gridOptions=gridOptions,
					height=grid_height,
					width='100%',
					data_return_mode=return_mode_value,
					update_mode=update_mode_value,
					fit_columns_on_grid_load=fit_columns_on_grid_load,
					allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
					enable_enterprise_modules=enable_enterprise_modules)
			else:
				st.error('Não há paletes sem selante utilizados na data selecionada')
		else:
			st.error('Não há paletes sem selante utilizados na data selecionada')
	else:
		st.error('Não há paletes sem selante utilizados na data selecionada')

	st.subheader('Selantes utilizadas na data selecionada')
	if df_selantes.shape[0] > 0:
		# selantes que possuem data de entrada e de saída
		selantes_filtradas = df_selantes.loc[(df_selantes['data_entrada'] != '-') ] 
		selantes_filtradas_s = df_selantes.loc[df_selantes['data_saida'] != '-']

		if (selantes_filtradas.shape[0] > 0) or (selantes_filtradas_s.shape[0] > 0):
			# converte os valores de string para datetime
			selantes_filtradas['data_entrada'] = pd.to_datetime(selantes_filtradas['data_entrada'])
			selantes_filtradas_s['data_saida'] = pd.to_datetime(selantes_filtradas_s['data_saida'])

			# filtra as selantes de acordo com a data
			filtro_selante = selantes_filtradas.loc[(selantes_filtradas['data_entrada'].dt.date == data_filtro)]
			filtro_selante_s = selantes_filtradas_s.loc[selantes_filtradas_s['data_saida'].dt.date == data_filtro]
			
			if (filtro_selante.shape[0] > 0) or (filtro_selante_s.shape[0] > 0):
				
				# combina os dados de entrada e os de saida
				resultado_c = filtro_selante.append(filtro_selante_s)

				# transforma as datas de volta em strings para facilitar a visualizacao
				resultado_c['data_entrada'] = pd.to_datetime(resultado_c['data_entrada'])
				resultado_c['data'] = pd.to_datetime(resultado_c['data'])

				# organiza pela data de entrada
				resultado_c = resultado_c.sort_values(by='data_entrada')

				# remove os duplicados filtrando pelo numero do lote
				resultado_c = resultado_c.drop_duplicates(subset='numero_lote')

				# organiza os dados para exibição
				#resultado_c['data_saida'] = resultado_c['data_saida'].apply(lambda x: '-' if x == '-' else x.dt.strftime("%H:%M %d/%m/%Y"))
				resultado_c['data'] = resultado_c['data'].dt.strftime("%d/%m/%Y")
				resultado_c['data_entrada'] = resultado_c['data_entrada'].dt.strftime("%H:%M %d/%m/%Y")
				
				#st.table(resultado_c)
				gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(120, resultado_c, 0, 0, True)
				response = AgGrid(
					resultado_c,
					gridOptions=gridOptions,
					height=grid_height,
					width='100%',
					data_return_mode=return_mode_value,
					update_mode=update_mode_value,
					fit_columns_on_grid_load=fit_columns_on_grid_load,
					allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
					enable_enterprise_modules=enable_enterprise_modules)
			else:
				st.error('Não há selantes utilizadas na data selecionada')
		else:
			st.error('Não há selantes utilizadas na data selecionada')
	else:
		st.error('Não há selantes utilizadas na data selecionada')

	st.subheader('Paletes com selante utilizados na data selecionada')
	if df_pal_com.shape[0] > 0:
		# bobinas que possuem data de entrada e de saída
		p_com_filtrado = df_pal_com.loc[df_pal_com['data_consumo'] != '-']

		if p_com_filtrado.shape[0] > 0:
			# converte os valores de string para datetime
			p_com_filtrado['data_consumo'] = pd.to_datetime(p_com_filtrado['data_consumo'])
			p_com_filtrado['data_gerado'] = pd.to_datetime(p_com_filtrado['data_gerado'])
			p_com_filtrado['data_estoque'] = pd.to_datetime(p_com_filtrado['data_estoque'])

			# filtra as bobinas de acordo com a data
			filtro_pal_com = p_com_filtrado.loc[(p_com_filtrado['data_consumo'].dt.date == data_filtro)]

			if filtro_pal_com.shape[0] > 0:
				# transforma as datas de volta em strings para facilitar a visualizacao
				filtro_pal_com['data_gerado'] = filtro_pal_com['data_gerado'].dt.strftime("%d/%m/%Y")
				filtro_pal_com['data_estoque'] = filtro_pal_com['data_estoque'].dt.strftime("%H:%M %d/%m/%Y")
				filtro_pal_com['data_consumo'] = filtro_pal_com['data_consumo'].dt.strftime("%H:%M %d/%m/%Y")

				#st.table(filtro_pal_com)
				gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(120, filtro_pal_com, 0, 0, True)
				response = AgGrid(
					filtro_pal_com,
					gridOptions=gridOptions,
					height=grid_height,
					width='100%',
					data_return_mode=return_mode_value,
					update_mode=update_mode_value,
					fit_columns_on_grid_load=fit_columns_on_grid_load,
					allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
					enable_enterprise_modules=enable_enterprise_modules)
			else:
				st.error('Não há paletes com selante utilizados na data selecionada')
		else:
			st.error('Não há paletes com selante utilizados na data selecionada')
	else:
		st.error('Não há paletes com selante utilizados na data selecionada')

if telas == 'Apontamento de código SAP':
	
	st.subheader('Apontamento de Código SAP')
	data_filtro = st.date_input('Selecione a data que deseja filtrar')
	st.subheader('Paletes sem selante')

	# seleciona as linhas que possuem data de estoque
	df_pal_sem_filtrado = df_pal_sem[df_pal_sem['data_estoque'] != '-']

	# transforma coluna no tipo datetime
	df_pal_sem_filtrado['data_estoque'] = pd.to_datetime(df_pal_sem_filtrado['data_estoque'])
	

	# filtra pela data selecionada
	if df_pal_sem_filtrado[df_pal_sem_filtrado['data_estoque'].dt.date == data_filtro].shape[0] > 0:
		
		# modifica o formato da data
		#df_pal_sem_filtrado['data_consumo'] = df_pal_sem_filtrado['data_consumo'].dt.strftime('%H:%M %d-%m-%Y')

		# escreve os valores filtrados
		#st.table(df_pal_sem_filtrado[df_pal_sem_filtrado['data_consumo'].dt.date == data_filtro])
		
		gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(150, df_pal_sem_filtrado[df_pal_sem_filtrado['data_estoque'].dt.date == data_filtro], 0, 0, True)
		response = AgGrid(
			df_pal_sem_filtrado[df_pal_sem_filtrado['data_estoque'].dt.date == data_filtro],
			gridOptions=gridOptions,
			height=grid_height,
			width='100%',
			data_return_mode=return_mode_value,
			update_mode=update_mode_value,
			fit_columns_on_grid_load=fit_columns_on_grid_load,
			allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
			enable_enterprise_modules=enable_enterprise_modules)

		# organiza as colunas
		valor, botao = st.beta_columns([9,1])

		# campo para incluir o codigo SAP do palete
		codigo_sap_sem = valor.text_input('Digite o código SAP para apontamento (sem selante)')

		# botao para modificar o codigo SAP
		modificar_sap_sem = botao.button('Apontamento de codigo SAP (sem selante)')

		if modificar_sap_sem:
			# flag para rodar novamente o script
			rerun = False

			# atribui o codigo sap aos paletes
			df_pal_sem.iloc[(df_pal_sem_filtrado['data_estoque'].dt.date == data_filtro).index, 6] = codigo_sap_sem

			# verifica as bobinas que pertecem os paletes
			unicos = list(df_pal_sem_filtrado.loc[df_pal_sem_filtrado['data_estoque'].dt.date == data_filtro, 'numero_OT'].unique())

			# itera sobre as bobinas
			for items in unicos:

				# prepara dados para escrever no banco
				dic_sap = {}
				dic_sap = df_bobinas.loc[(df_bobinas['numero_OT'] == items)].to_dict('records')

				# Transforma dados do formulário em um dicionário
				keys_values = dic_sap[0].items()
				new_sap = {str(key): str(value) for key, value in keys_values}
				documento_sap = new_sap['numero_OT']

				# escreve o dataframe dos paletes na selante para escrita em banco (não altera valor, mas escreve para não perder os dados)
				new_sap['Paletes'] = df_pal_sem.loc[(df_pal_sem['numero_OT'] == items)].to_csv()

				# Armazena no banco as alteracoes da bobina
				try:
					doc_ref = db.collection("Bobina").document(documento_sap)
					doc_ref.set(new_sap)
					st.success('Modificação armazenada com sucesso!')
					rerun = True
				except:
					st.error('Falha ao armazenar modificação, tente novamente ou entre em contato com suporte!')
					caching.clear_cache()

			# comando para rodar novament o script
			if rerun:
				st.experimental_rerun()

			st.write(df_pal_sem)
	else:
		st.error('Não há paletes para serem apontados para data selecionada')

	st.subheader('Paletes com selante')

	# seleciona as linhas que possuem data de estoque
	df_pal_com_filtrado = df_pal_com[df_pal_com['data_estoque'] != '-']

	# transforma coluna no tipo datetime
	df_pal_com_filtrado['data_estoque'] = pd.to_datetime(df_pal_com_filtrado['data_estoque'])
	

	# filtra pela data selecionada
	if df_pal_com_filtrado[df_pal_com_filtrado['data_estoque'].dt.date == data_filtro].shape[0] > 0:

		# modifica o formato da data
		#df_pal_com_filtrado['data_consumo'] = df_pal_com_filtrado['data_consumo'].dt.strftime('%H:%M %d-%m-%Y')

		# escreve os valores filtrados
		#st.table(df_pal_com_filtrado[df_pal_com_filtrado['data_consumo'].dt.date == data_filtro])

		gridOptions, grid_height, return_mode_value, update_mode_value, fit_columns_on_grid_load, enable_enterprise_modules = config_grid(150, df_pal_com_filtrado[df_pal_com_filtrado['data_estoque'].dt.date == data_filtro], 0, 0, True)
		response = AgGrid(
			df_pal_com_filtrado[df_pal_com_filtrado['data_estoque'].dt.date == data_filtro],
			gridOptions=gridOptions,
			height=grid_height,
			width='100%',
			data_return_mode=return_mode_value,
			update_mode=update_mode_value,
			fit_columns_on_grid_load=fit_columns_on_grid_load,
			allow_unsafe_jscode=False,  # Set it to True to allow jsfunction to be injected
			enable_enterprise_modules=enable_enterprise_modules)

		# organiza as colunas
		valor, botao = st.beta_columns([9,1])

		# campo para incluir o codigo SAP do palete
		codigo_sap_com = valor.text_input('Digite o código SAP para apontamento (com selante)')

		# botao para modificar o codigo SAP
		modificar_sap_com = botao.button('Apontamento de codigo SAP (com selante)')

		if modificar_sap_com:
			# flag para rodar novamente o script
			rerun = False

			# atribui o codigo sap aos paletes
			df_pal_com.iloc[(df_pal_com_filtrado['data_estoque'].dt.date == data_filtro).index, 4] = codigo_sap_com

			# verifica as bobinas que pertecem os paletes
			unicos = list(df_pal_com_filtrado.loc[df_pal_com_filtrado['data_estoque'].dt.date == data_filtro, 'numero_lote'].unique())

			# itera sobre as bobinas
			for items in unicos:

				# prepara dados para escrever no banco
				dic_sap = {}
				dic_sap = df_selantes.loc[(df_selantes['numero_lote'] == items)].to_dict('records')

				# Transforma dados do formulário em um dicionário
				keys_values = dic_sap[0].items()
				new_sap = {str(key): str(value) for key, value in keys_values}
				documento_sap = new_sap['numero_lote']

				# escreve o dataframe dos paletes na selante para escrita em banco (não altera valor, mas escreve para não perder os dados)
				new_sap['Paletes'] = df_pal_com.loc[(df_pal_com['numero_lote'] == items)].to_csv()

				# Armazena no banco as alteracoes da bobina
				try:
					doc_ref = db.collection("Selante").document(documento_sap)
					doc_ref.set(new_sap)
					st.success('Modificação armazenada com sucesso!')
					rerun = True
				except:
					st.error('Falha ao armazenar modificação, tente novamente ou entre em contato com suporte!')
					caching.clear_cache()

			# comando para rodar novament o script
			if rerun:
				st.experimental_rerun()

			st.write(df_pal_com)
	else:
		st.error('Não há paletes para serem apontados para data selecionada')

# botao para teste
reset = st.button('Reset')

if reset:
	caching.clear_cache()

